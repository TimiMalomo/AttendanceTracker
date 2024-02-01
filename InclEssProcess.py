import pandas as pd
import openpyxl
from pathlib import Path
from datetime import datetime

def is_valid_date(date_str):
    try:
        pd.to_datetime(date_str)
        return True
    except (ValueError, OverflowError):
        return False

def update_attendance_tracker(ess_summaries_file, tracker_xlsx):
    # Read data from ESSummaries.txt
    ess_data = pd.read_csv(ess_summaries_file, delimiter="\t")

    # Load the AttendanceTracker workbook
    wb = openpyxl.load_workbook(tracker_xlsx)

    # Iterate over each row in the ESSummaries
    for _, row in ess_data.iterrows():
        initiator = row['Initiator']
        start_date_str = row['Start Date']
        if not is_valid_date(start_date_str):
            print(f"Invalid date format for {initiator}: {start_date_str}")
            continue
        start_date = pd.to_datetime(start_date_str).date()  # Convert to datetime

        att_abs_type = row['Att./abs. type text']
        submit_date = row['Submit Date']
        approver = row['Approver']
        end_date = row['End Date']
        absence_hrs = row['Absence hrs']

        # Check if the sheet for the Initiator exists
        if initiator in wb.sheetnames:
            sheet = wb[initiator]

            # Flag to check if changes were made
            changes_made = False

            # Iterate over the rows in the sheet
            for row in sheet.iter_rows(min_row=2):  # Assuming row 1 is the header
                if row[0].value and pd.to_datetime(row[0].value).date() == start_date:
                    # Update columns B, J, N, and V
                    row[1].value = att_abs_type  # Column B
                    row[9].value = att_abs_type  # Column J
                    row[13].value = 'Y'         # Column N
                    row[21].value = f"{submit_date}+{approver}+{start_date}+{end_date}+{absence_hrs}"  # Column V
                    changes_made = True
            
            # Print a message if changes were made
            if changes_made:
                print(f"Updated sheet for {initiator}")
            else:
                print(f"No matching dates found for {initiator}")

        else:
            print(f"Sheet not found for {initiator}")

    # Save the workbook
    wb.save(tracker_xlsx)

# Call the function
update_attendance_tracker('ESSummaries.txt', 'AttendanceTracker.xlsx')
