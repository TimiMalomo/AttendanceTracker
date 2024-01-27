import openpyxl

def process_summaries(input_txt, input_xlsx):
    # Open the Excel workbook
    workbook = openpyxl.load_workbook(input_xlsx)

    # Read data from Summaries.txt
    with open(input_txt, 'r') as file:
        next(file)  # Skip the first line (header)
        for line in file:
            # Split the line into components
            employee_name, report_date, activity_name, activity_start, activity_end, duration = line.strip().split('\t')

            # Check if the sheet with the employee's name exists in the workbook
            if employee_name in workbook.sheetnames:
                sheet = workbook[employee_name]
                # Find the next empty row after row 35
                row = 36
                while sheet[f'A{row}'].value is not None:
                    row += 1

                # Write data to the cells, handling merged cells
                sheet[f'A{row}'] = report_date
                sheet[f'C{row}'] = duration
                # Handle merged cells for column E
                e_cell_written = False
                for range_ in sheet.merged_cells.ranges:
                    if range_.min_row <= row <= range_.max_row and range_.min_col <= 5 <= range_.max_col:
                        top_left_cell = sheet.cell(row=range_.min_row, column=range_.min_col)
                        top_left_cell.value = activity_name
                        e_cell_written = True
                        break
                if not e_cell_written:
                    sheet.cell(row=row, column=5).value = activity_name  # Column E is 5th column

                # Write to column Q (17th column)
                sheet.cell(row=row, column=17).value = f'{activity_name} {activity_start}-{activity_end}'
            else:
                print(f"Sheet for {employee_name} not found in workbook.")

    # Save the updated workbook
    workbook.save(input_xlsx)

# Call the function with file names
process_summaries('Summaries.txt', 'AttendanceTracker.xlsx')
