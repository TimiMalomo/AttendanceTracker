import openpyxl

# Open the workbook
wb = openpyxl.load_workbook('AttendanceTracker.xlsx')

# Function to match and update records in the sheet, including appending information to column Q
def update_sheet(sheet, start_date, att_abs_type, submit_date, approver, absence_hrs):
    for row in sheet.iter_rows(min_row=2, max_col=17):  # Ensure max_col includes column Q
        cell_date = row[0].value  # Dates are in column A
        if cell_date == start_date:
            row[9].value = att_abs_type  # Update column J
            row[13].value = 'Y'  # Update column N
            
            # Append information to column Q
            existing_info = row[16].value if row[16].value is not None else ""  # Existing content in Q
            additional_info = f"{existing_info} Submit Date: {submit_date}, Approver: {approver}, Hours Submitted: {absence_hrs}"
            row[16].value = additional_info  # Update column Q with the new string
            return True
    return False

# Read the ESSSummaries.txt file
with open('ESSummaries.txt', 'r') as file:
    next(file)  # Skip the header line
    for line in file:
        components = line.strip().split('\t')
        if len(components) < 7:
            continue
        
        submit_date, initiator, approver, att_abs_type, start_date, _, absence_hrs = components
        
        # Use the initiator directly as the sheet name
        sheet_name = initiator
        
        # Find the corresponding sheet and update it
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            update_sheet(sheet, start_date, att_abs_type, submit_date, approver, absence_hrs)

# Save the workbook with changes
wb.save('attendancetracker_updated.xlsx')
